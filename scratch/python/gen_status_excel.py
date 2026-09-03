"""Generate a manager-facing Excel status sheet of completed masters/pages."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"D:\project\CareFusions\CareFusions_Masters_Status.xlsx"

# (Menu, Master Page, API endpoint, Status, Key capabilities)
ROWS = [
    # Radiology
    ("Radiology Masters", "Radiology Service Master", "/api/v1/radiology-services", "Completed", "CRUD, auto-code, name-unique, soft delete, field limits"),
    ("Radiology Masters", "Equipment Master", "/api/v1/equipment", "Completed", "CRUD, auto-code, uniqueness, validation, limits"),
    # Billing
    ("Billing Masters", "Service Master", "/api/v1/services", "Completed", "CRUD, auto-code, price/tax validation, limits"),
    ("Billing Masters", "Tax (GST) Master", "/api/v1/taxes", "Completed", "CRUD, rate validation, SP-enforced uniqueness"),
    ("Billing Masters", "Payment Mode Master", "/api/v1/payment-modes", "Completed", "CRUD, auto-code, uniqueness, limits"),
    # Insurance
    ("Insurance Masters", "Insurance Provider Master", "/api/v1/insurance-providers", "Completed", "CRUD, auto-code, contact/email validation"),
    ("Insurance Masters", "TPA Master", "/api/v1/tpas", "Completed", "CRUD, auto-code, uniqueness, limits"),
    # Purchase & Inventory
    ("Purchase & Inventory", "Vendor Master", "/api/v1/vendors", "Completed", "CRUD, GST/contact validation, limits"),
    ("Purchase & Inventory", "Category Master", "/api/v1/categories", "Completed", "CRUD, auto-code, uniqueness"),
    ("Purchase & Inventory", "Sub-Category Master", "/api/v1/sub-categories", "Completed", "CRUD, cascades from Category"),
    ("Purchase & Inventory", "UOM Master", "/api/v1/uoms", "Completed", "CRUD, auto-code, uniqueness"),
    ("Purchase & Inventory", "Item Master", "/api/v1/items", "Completed", "CRUD, live Category/UOM/Brand lookups"),
    ("Purchase & Inventory", "Brand Master", "/api/v1/brands", "Completed", "CRUD, auto-code, uniqueness"),
    ("Purchase & Inventory", "Manufacturer Master", "/api/v1/manufacturers", "Completed", "CRUD, auto-code, uniqueness"),
    ("Purchase & Inventory", "Store / Warehouse Master", "/api/v1/stores", "Completed", "CRUD, live Hospital/Branch lookups"),
    # Financial
    ("Financial Masters", "Chart of Accounts", "/api/v1/coa", "Completed", "CRUD, account-type validation, uniqueness"),
    ("Financial Masters", "Cost Center Master", "/api/v1/cost-centers", "Completed", "CRUD, auto-code, uniqueness"),
    ("Financial Masters", "Profit Center Master", "/api/v1/profit-centers", "Completed", "CRUD, auto-code, uniqueness"),
    ("Financial Masters", "Payment Terms Master", "/api/v1/payment-terms", "Completed", "CRUD, days validation, limits"),
    ("Financial Masters", "Currency Master", "/api/v1/currencies", "Completed", "CRUD, code/symbol validation, uniqueness"),
    ("Financial Masters", "Financial Year Master", "/api/v1/financial-years", "Completed", "CRUD, date-range validation"),
    ("Financial Masters", "Bank Master", "/api/v1/banks", "Completed", "CRUD, IFSC/account validation, limits"),
    ("Financial Masters", "Cash Counter Master", "/api/v1/cash-counters", "Completed", "CRUD, live Hospital/Branch, opening<=limit rule"),
    # Security
    ("Security Masters", "Role Master", "/api/v1/roles", "Completed", "CRUD, single-default rule, live user-count, permission flags"),
    ("Security Masters", "User Master", "/api/v1/users", "Completed", "CRUD, PBKDF2 password hashing, global role + permissions, Hospital->Branch cascade, one-active-account/employee"),
    # Notification
    ("Notification Masters", "SMS Template Master", "/api/v1/sms-templates", "Completed", "CRUD, name-unique, variable insert, 1000-char cap"),
    ("Notification Masters", "Email Template Master", "/api/v1/email-templates", "Completed", "CRUD, HTML body, attachment rule (type required if allowed)"),
    ("Notification Masters", "WhatsApp Template Master", "/api/v1/whatsapp-templates", "Completed", "CRUD, Meta template-ID format+uniqueness, language enum"),
    ("Notification Masters", "Push Notification Template", "/api/v1/push-templates", "Completed", "CRUD, priority enum, deep-link URL validation"),
    ("Notification Masters", "Reminder Rule Master", "/api/v1/reminder-rules", "Completed", "CRUD, channel enum, >=1 recipient, repeat/retry rules"),
    # Audit
    ("Audit Logs", "Audit Log Master", "/api/v1/audit-logs", "Completed", "Read-only + append-only (immutable), search/filters, PDF/Excel export"),
    # Core lookups
    ("Core / Setup", "Hospital Master", "/api/v1/hospitals", "Available", "Integrated as live lookup across masters"),
    ("Core / Setup", "Branch Master", "/api/v1/branches", "Available", "Integrated as live lookup (Hospital-linked)"),
    ("Core / Setup", "Department Master", "/api/v1/departments", "Available", "Integrated as live lookup across masters"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Masters Status"

# ── Palette ──
NAVY   = "1F3A5F"
LBLUE  = "DCE6F2"
GREEN  = "2E7D32"
GREENF = "E6F4EA"
GREY   = "6B7280"
WHITE  = "FFFFFF"
thin = Side(style="thin", color="D0D5DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Title ──
ws.merge_cells("A1:F1")
t = ws["A1"]
t.value = "CareFusions HMS — Admin Masters: Completion Status"
t.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=NAVY)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:F2")
st = ws["A2"]
completed = sum(1 for r in ROWS if r[3] == "Completed")
st.value = (f"{completed} master pages delivered full-stack (backend API + database + frontend, tested).  "
            f"Report date: 2026-07-31")
st.font = Font(name="Calibri", size=10, italic=True, color="374151")
st.fill = PatternFill("solid", fgColor=LBLUE)
st.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 20

# ── Header row ──
HEAD = ["S.No", "Menu / Module", "Master Page", "API Endpoint", "Status", "Key Capabilities"]
hrow = 3
for c, h in enumerate(HEAD, start=1):
    cell = ws.cell(row=hrow, column=c, value=h)
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = border
ws.row_dimensions[hrow].height = 22

# ── Data rows ──
r = hrow + 1
for i, (menu, page, api, statusv, caps) in enumerate(ROWS, start=1):
    vals = [i, menu, page, api, statusv, caps]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=(c == 6))
        if i % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F7F9FC")
        if c == 4:
            cell.font = Font(name="Consolas", size=9, color="374151")
        if c == 5:
            if statusv == "Completed":
                cell.font = Font(bold=True, color=GREEN)
                cell.fill = PatternFill("solid", fgColor=GREENF)
            else:
                cell.font = Font(bold=True, color=GREY)
                cell.fill = PatternFill("solid", fgColor="F3F4F6")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

# ── Column widths ──
widths = [6, 22, 30, 30, 13, 62]
for c, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A4"
ws.sheet_view.showGridLines = False

wb.save(OUT)
print("Saved:", OUT)
print("Rows:", len(ROWS), "| Completed:", completed)
